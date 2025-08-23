import configparser
import hashlib
import json
import time
from pathlib import Path
from openpyxl import load_workbook

start_time = time.time()
root_dir = Path(__file__).parent.resolve()

def calculate_hash(file_path):
    if not file_path.exists():
        return None
    hasher = hashlib.md5()
    with file_path.open("rb") as f:
        while chunk := f.read(4096):
            hasher.update(chunk)
    return hasher.hexdigest()

config = configparser.ConfigParser()
config.read(root_dir / "config.ini")

total_files = sum(int(config[s]["FileNum"]) for s in config.sections() if s.startswith("Dir"))
processed = 0

for section in config.sections():
    if section.startswith("Dir"):
        dir_rel_path = config[section]["Directory"].lstrip("/\\")
        dir_path = root_dir / dir_rel_path

        if not dir_path.is_dir():
            print(f"Skipping missing directory: {dir_path}")
            continue

        file_num = int(config[section]["FileNum"])

        for i in range(file_num):
            processed += 1
            elapsed = time.time() - start_time
            print(f"Processing file {processed} of {total_files} | Elapsed: {elapsed:.2f}s")
            file_name = config[section][f"File{i}"]
            file_path = dir_path / file_name
            hash_file_path = file_path.with_suffix(file_path.suffix + ".hash")
            encoding = config[section].get(f"EncodingFile{i}", "cp1252")

            new_hash = calculate_hash(file_path)
            if new_hash is None:
                print(f"File not found: {file_path}")
                continue

            old_data = {}
            if hash_file_path.exists():
                try:
                    with hash_file_path.open("r", encoding="utf-8") as hf:
                        old_data = json.load(hf)
                except json.JSONDecodeError:
                    old_data = {}

            old_hash = old_data.get("hash", "")
            old_txt_files = set(old_data.get("txt_files", []))

            export_required = (new_hash != old_hash) or any(
                not (dir_path / txt_file).exists() for txt_file in old_txt_files
            )

            if export_required:
                wb = load_workbook(file_path, data_only=True, read_only=True)
                new_txt_files = []

                for ws in wb.worksheets:
                    lines = []
                    first_row = True
                    for row in ws.iter_rows(values_only=True):
                        line = "\t".join(
                            f"{x:.15g}" if isinstance(x, float)
                            else (str(x) if x is not None else "\t")
                            for x in row
                        )
                        if first_row:
                            line += "\t"
                            first_row = False
                        lines.append(line + "\n")

                    txt_file_path = dir_path / f"{ws.title}.txt"
                    txt_file_path.write_text("".join(lines), encoding=encoding, errors="replace")
                    new_txt_files.append(f"{ws.title}.txt")

                with hash_file_path.open("w", encoding="utf-8") as hf:
                    json.dump({"hash": new_hash, "txt_files": new_txt_files}, hf)

                print(f"Processing: {file_path} with encoding {encoding}")
            else:
                print(f"Skipped (unchanged): {file_path}")